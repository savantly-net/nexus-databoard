from enum import Enum
import os
import random
import shutil
from zipfile import ZipFile
import streamlit as st
import pandas as pd
from pandas import DataFrame
import nexus_databoard
from nexus_databoard.components import table, download
from IPython.display import HTML
from openpyxl import load_workbook
from pathlib import Path

project_directory = os.path.dirname(os.path.realpath(__file__)) + "/.."

class EMPTY_FILE_TYPE(Enum):
    custom_remort = f"{project_directory}/templates/custom_remort.xlsx"

session_id = None

def create_new_session_id():
    st.session_state["session_id"] = random.randint(1000, 1000000)

# create a unique session id
if "session_id" not in st.session_state:
    create_new_session_id()
session_id = st.session_state["session_id"]

export_directory = f"{project_directory}/temp/"

# get the file path
def get_temp_file_path(file_name):
    if not os.path.exists(export_directory):
        os.makedirs(export_directory)
    return os.path.join(export_directory, file_name)

# exported file names
export_custom_remort_file_name = f"{session_id}_exported_custom_remort.xlsx"
export_custom_remort_file_path = get_temp_file_path(export_custom_remort_file_name)


# Get file bytes
def get_file_bytes(file_path):
    with open(file_path, "rb") as f:
        bytes = f.read()
    return bytes

# Create an empty xlsx file
def create_empty_xls_file(file_path, template_path: EMPTY_FILE_TYPE):
    # copy empty file to the export directory
    shutil.copyfile(template_path.value, file_path)

# Write bytes to xls file
def write_sheet_to_xls_file(df: DataFrame, sheet_name, file_path, template_path: EMPTY_FILE_TYPE):
    # check if xls file exists, and create an empty one if it doesn't
    if not os.path.exists(file_path):
        create_empty_xls_file(file_path, template_path=template_path)

    # load the workbook
    ExcelWorkbook = load_workbook(file_path)

    # check if the sheet already exists in the workbook
    if sheet_name in ExcelWorkbook.sheetnames:
        # get the sheet
        sheet = ExcelWorkbook[sheet_name]
        # delete everything after the header row
        sheet.delete_rows(2, sheet.max_row)
        # convert DataFrame to list of lists (values)
        data = df.values.tolist()
        # append the data to the existing sheet
        for row_data in data:
            sheet.append(row_data)
    else:
        # create a new sheet
        ExcelWorkbook.create_sheet(title=sheet_name)
        # get the sheet
        sheet = ExcelWorkbook[sheet_name]
        # create header row
        header = df.columns.tolist()
        sheet.append(header)
        # convert DataFrame to list of lists (values)
        data = df.values.tolist()
        # write the data to the new sheet
        for row_data in data:
            sheet.append(row_data)

    # save the changes
    ExcelWorkbook.save(file_path)

def draw_table(df):
    st.dataframe(df)

def show_download_button(label, file_path, file_name, mime="application/vnd.ms-excel"):
    st.download_button(
        key=file_name,
        label=label,
        data=get_file_bytes(file_path),
        file_name=file_name,
        mime=mime
    )


def extract_custom_remort(file):
    st.write("custom_remort")
    df = pd.read_excel(file, sheet_name='custom_remort', header=1)

    # delete the first column
    del df[df.columns[0]]
    
    # delete the extra columns at the end
    del df[df.columns[10]]
    del df[df.columns[9]]
    del df[df.columns[8]]

    # Don't include rows where the total is 0
    df = df[df.iloc[:, 4] != "0"]
    
    # Create the columns for the exported sheet
    df_exported = pd.DataFrame()
    df_exported['Name'] = "" 
    
    df_exported = df_exported[~df_exported['Description'].isin(['-', 'non'])]
    df_exported = df_exported.dropna(subset=['Description'])

    sheet_name = "custom_remort's"  # New sheet name
    write_sheet_to_xls_file(df_exported, sheet_name, export_custom_remort_file_path, template_path=EMPTY_FILE_TYPE.custom_remort)
    draw_table(df_exported)


        
def parse_file(file):
    if file is not None:

        st.subheader("custom_remorts")
        extract_custom_remort(file)
        # Add download button
        show_download_button("Download custom_remort", export_custom_remort_file_path, export_custom_remort_file_name)

################## Main ##################
st.title("Savantly Custom Report")
st.write("Upload a file to generate report")

def reset_session_id():
    create_new_session_id()

file = st.file_uploader("Upload a file", type=["xls", "xlsx"], on_change=reset_session_id)

if file is not None:
    parse_file(file)

    st.subheader("Download All as Zip")

    # zip the files
    zip_file_name = "exported_files.zip"
    zip_file_path = get_temp_file_path(zip_file_name)
    with ZipFile(zip_file_path, 'w') as zip:
        zip.write(export_custom_remort_file_path, export_custom_remort_file_name)

    
    # Add download button
    show_download_button("Download Zip", zip_file_path, zip_file_name, mime="application/zip")